import openpyxl

# Open the two Excel files
wb1 = openpyxl.load_workbook('file1.xlsx')
wb2 = openpyxl.load_workbook('file2.xlsx')

# Get the sheet names in each workbook
sheet_names1 = wb1.sheetnames
sheet_names2 = wb2.sheetnames

# Check if the sheet names are the same in both workbooks
if sheet_names1 != sheet_names2:
    print('The sheet names are different')

# If the sheet names are the same, compare the sheet data
else:
    # Iterate through each sheet in the workbook
    for sheet_name in sheet_names1:
        # Get the sheet objects
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        # Compare the sheet data cell by cell
        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                cell1 = sheet1.cell(row, col).value
                cell2 = sheet2.cell(row, col).value
                if cell1 != cell2:
                    print(f'Difference found at cell {row},{col}')
